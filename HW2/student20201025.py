#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']
total = list()

row_id = 1;
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		total.append(sum_v)
	row_id += 1

total.sort(reverse = True)

row_id = 1
for row in ws:
	if row_id != 1:
		if ws.cell(row = row_id, column = 7).value >= total[10]:
			ws.cell(row = row_id, column = 8).value = "A+"
		elif total[22] <=  ws.cell(row = row_id, column = 7).value <= total[11]:
			ws.cell(row = row_id, column = 8).value = "A"
		elif total[36] <= ws.cell(row = row_id, column = 7).value <= total[23]:
			ws.cell(row = row_id, column = 8).value = "B+"
		elif total[51] <= ws.cell(row = row_id, column = 7).value <= total[37]:
                        ws.cell(row = row_id, column = 8).value = "B"
		elif total[62] <= ws.cell(row = row_id, column = 7).value <= total[52]:
                        ws.cell(row = row_id, column = 8).value = "C+"
		elif total[73] <= ws.cell(row = row_id, column = 7).value <= total[63]:
                        ws.cell(row = row_id, column = 8).value = "C"
	row_id += 1

wb.save("student.xlsx")

