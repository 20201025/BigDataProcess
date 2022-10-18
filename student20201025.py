#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']
total = list()

row_id = 1;
for row in ws:
    if row_id != 1:
        sum_v = ws.cell(row = row_id, column = 3).value * 0.3
        sum_v += ws.cell(row = row_id, column = 4).value * 0.35
        sum_v += ws.cell(row = row_id, column = 5).value * 0.34
        sum_v += ws.cell(row = row_id, column = 6).value
        ws.cell(row = row_id, column = 7).value = sum_v
        total.append(sum_v)
    row_id += 1

total.sort(reverse = True)
grade = [0, 0, 0, 0, 0, 0]

row_id = 1
for row in ws:
    if row_id != 1:
        if ws.cell(row = row_id, column = 7).value >= total[10]:
            ws.cell(row = row_id, column = 8).value = "A+"
            grade[0] += 1

        elif total[21] <=  ws.cell(row = row_id, column = 7).value <= total[11]:
            ws.cell(row = row_id, column = 8).value = "A"
            grade[1] += 1

        elif total[35] <= ws.cell(row = row_id, column = 7).value <= total[22]:
            ws.cell(row = row_id, column = 8).value = "B+"
            grade[2] += 1

        elif total[50] <= ws.cell(row = row_id, column = 7).value <= total[36]:
            ws.cell(row = row_id, column = 8).value = "B"
            grade[3]+=1

        elif total[61] <= ws.cell(row = row_id, column = 7).value <= total[51]:
            ws.cell(row = row_id, column = 8).value = "C+"
            grade[4]+=1

        elif total[72] <= ws.cell(row = row_id, column = 7).value <= total[62]:
            ws.cell(row = row_id, column = 8).value = "C"
            grade[5]+=1

        elif total[0] == total[73]:
            ws.cell(row = row_id, column = 8).value = "C"
    row_id += 1


print(grade)
wb.save("student.xlsx")
